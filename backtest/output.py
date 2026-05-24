"""
输出层 — 控制台格式化打印 + Excel 导出
"""

import os
from typing import List

from backtest.strategy_base import ScreenResult
from core.config import OUTPUT_DIR, OUTPUT_FILENAME_FORMAT


# ═════════════════════════════════════════
# 控制台输出
# ═════════════════════════════════════════

def print_table(results: List[ScreenResult], limit: int = 30) -> None:
    """格式化打印筛选结果到控制台"""
    if not results:
        print("  （无通过筛选的股票）")
        return

    total = len(results)
    show = results[:limit]
    remaining = total - len(show)

    print(f"\n  通过筛选: {total} 只", end="")
    if remaining > 0:
        print(f"（显示前 {len(show)} 只，共 {total} 只）")
    else:
        print()
    print()

    # 表头
    header = (
        f"{'序号':<4} {'代码':<8} {'名称':<10} "
        f"{'价格':<8} {'涨跌幅':<8}"
    )
    print("  " + header)
    print("  " + "-" * len(header))

    for i, r in enumerate(show, 1):
        pct_str = f"{r.pct_change:+.2f}%" if r.pct_change else "N/A"
        price_str = f"{r.price:.2f}" if r.price else "N/A"
        print(
            f"  {i:<4} {r.code:<8} {r.name:<10} "
            f"{price_str:<8} {pct_str:<8}"
        )

        # 打印各条件明细
        for fr in r.filter_results:
            status = "✓" if fr.passed else "✗"
            detail = f" ({fr.detail})" if fr.detail and not fr.passed else ""
            print(f"       {status} {fr.name}{detail}")

        if r.signal_desc:
            print(f"       信号: {r.signal_desc}")
        print()

    if remaining > 0:
        print(f"  ... 还有 {remaining} 只未显示")


# ═════════════════════════════════════════
# Excel 导出
# ═════════════════════════════════════════

def to_excel(results: List[ScreenResult], filename: str = None) -> str:
    """
    导出筛选结果到 Excel。
    返回文件路径。
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment
    except ImportError:
        raise ImportError("请先安装 openpyxl: pip install openpyxl")

    if not results:
        print("  （无数据，跳过 Excel 导出）")
        return ""

    # 确保输出目录存在
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # 生成文件名
    if not filename:
        from datetime import datetime
        date_str = datetime.now().strftime("%Y%m%d_%H%M")
        filename = OUTPUT_FILENAME_FORMAT.replace("{date}", date_str)

    filepath = os.path.join(OUTPUT_DIR, filename)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "筛选结果"

    # 收集所有条件名称（按第一个结果的顺序）
    all_filter_names = []
    if results[0].filter_results:
        all_filter_names = [fr.name for fr in results[0].filter_results]

    # 表头
    headers = ["序号", "代码", "名称", "价格", "涨跌幅"] + all_filter_names + ["信号描述"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        cell.alignment = Alignment(horizontal="center")

    # 数据行
    for row_idx, r in enumerate(results, 2):
        ws.cell(row=row_idx, column=1, value=row_idx - 1)
        ws.cell(row=row_idx, column=2, value=r.code)
        ws.cell(row=row_idx, column=3, value=r.name)
        ws.cell(row=row_idx, column=4, value=r.price)
        ws.cell(row=row_idx, column=5, value=r.pct_change)

        # 各条件结果
        fr_map = {fr.name: fr for fr in r.filter_results}
        for col_offset, fn in enumerate(all_filter_names):
            fr = fr_map.get(fn)
            if fr:
                cell = ws.cell(row=row_idx, column=6 + col_offset)
                if fr.passed:
                    cell.value = "✓"
                    cell.font = Font(color="00B050")
                else:
                    cell.value = f"✗ {fr.detail}"
                    cell.font = Font(color="FF0000")
                cell.alignment = Alignment(horizontal="center")

        # 信号描述
        ws.cell(
            row=row_idx,
            column=6 + len(all_filter_names),
            value=r.signal_desc,
        )

    # 冻结首行 + 列宽
    ws.freeze_panes = "A2"
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    wb.save(filepath)
    print(f"\n  Excel 已导出: {filepath}")
    return filepath
